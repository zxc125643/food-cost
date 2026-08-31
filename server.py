#!/usr/bin/env python3
"""通用食品成本计算器 - 后端服务
物料价格全局共享 + 产品配方独立配置
端口 8092（与月饼计算器 8090 完全独立）
"""
import http.server
import socketserver
import json
import os
import re
import io
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PORT = 8092
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MATERIALS_FILE = os.path.join(BASE_DIR, "materials.json")
PRODUCTS_FILE = os.path.join(BASE_DIR, "products.json")


# ---------- 数据读写 ----------

def load_json(path, default):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def save_json(path, data):
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def load_materials():
    return load_json(MATERIALS_FILE, {})


def load_products():
    return load_json(PRODUCTS_FILE, {})


# ---------- Excel 导出（公式联动） ----------

def build_excel(materials, product):
    """生成带公式联动的 Excel：改价格/用量 → 成本自动重算"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成本计算"

    # 样式
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    input_fill = PatternFill("solid", fgColor="FFF2CC")   # 黄色=可改
    calc_fill = PatternFill("solid", fgColor="E2EFDA")    # 绿色=自动算
    subtotal_fill = PatternFill("solid", fgColor="DDEBF7")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center")

    p_name = product.get("name", "产品")
    specs = product.get("specs", [])
    usage = product.get("usage", {})

    # 标题
    ws.merge_cells("A1:" + get_col_letter(len(specs) + 2) + "1")
    ws["A1"] = f"{p_name} 单个成本计算表（改黄色格，绿色自动重算）"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # ===== 区域1：物料价格（全局共享）=====
    ws["A3"] = "① 物料价格（全局共享，所有产品通用）"
    ws["A3"].font = Font(bold=True, size=12)

    mat_keys = [k for k in materials.keys() if not k.startswith("_")]
    # 只保留该产品用到的物料
    used_mats = set()
    for spec in specs:
        for mk in usage.get(spec, {}).keys():
            used_mats.add(mk)
    mat_keys = [k for k in mat_keys if k in used_mats]

    row = 4
    headers = ["物料", "购买单位", "每份价格(元)", "每份净含量(克)", "折算单价(元/克)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = center

    mat_row = {}  # mat_key -> excel row
    for mk in mat_keys:
        row += 1
        m = materials.get(mk, {})
        ws.cell(row=row, column=1, value=m.get("name", mk)).border = border
        uc = ws.cell(row=row, column=2, value=m.get("unit", ""))
        uc.fill = input_fill; uc.border = border; uc.alignment = center
        pc = ws.cell(row=row, column=3, value=num(m.get("price")))
        pc.fill = input_fill; pc.border = border; pc.alignment = center
        wc = ws.cell(row=row, column=4, value=num(m.get("weight")))
        wc.fill = input_fill; wc.border = border; wc.alignment = center
        fc = ws.cell(row=row, column=5,
                     value=f'=IF(AND(C{row}<>"",D{row}<>""),C{row}/D{row},"")')
        fc.fill = calc_fill; fc.border = border; fc.alignment = center
        fc.number_format = "0.0000"
        mat_row[mk] = row

    # ===== 区域2：用量 + 成本矩阵 =====
    row += 2
    ws.cell(row=row, column=1, value=f"② {p_name} 各规格用量与成本").font = Font(bold=True, size=12)
    row += 1
    header_row = row
    ws.cell(row=row, column=1, value="物料").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).border = border
    ws.cell(row=row, column=1).alignment = center
    for i, spec in enumerate(specs):
        c = ws.cell(row=row, column=2 + i, value=f"{spec} 用量(克)")
        c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = center
        c2 = ws.cell(row=row, column=2 + i + len(specs), value=f"{spec} 成本(元)")
        c2.font = header_font; c2.fill = header_fill; c2.border = border; c2.alignment = center

    # 用量行
    first_data_row = row + 1
    for mk in mat_keys:
        row += 1
        m = materials.get(mk, {})
        ws.cell(row=row, column=1, value=m.get("name", mk)).border = border
        for i, spec in enumerate(specs):
            u = usage.get(spec, {}).get(mk, "")
            uc = ws.cell(row=row, column=2 + i, value=num(u))
            uc.fill = input_fill; uc.border = border; uc.alignment = center
            # 成本 = 用量 × 折算单价
            cost_col = 2 + i + len(specs)
            cc = ws.cell(row=row, column=cost_col,
                         value=f'=IF(OR({col_letter(2+i)}{row}="",E{mat_row[mk]}=""),"",'
                               f'{col_letter(2+i)}{row}*E{mat_row[mk]})')
            cc.fill = calc_fill; cc.border = border; cc.alignment = center
            cc.number_format = "0.0000"

    # 合计行
    row += 1
    total_row = row
    ws.cell(row=row, column=1, value="单个合计").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=1).border = border
    for i, spec in enumerate(specs):
        cost_col = 2 + i + len(specs)
        cl = col_letter(cost_col)
        sc = ws.cell(row=row, column=cost_col,
                     value=f'=SUM({cl}{first_data_row}:{cl}{row-1})')
        sc.font = Font(bold=True)
        sc.fill = subtotal_fill; sc.border = border; sc.alignment = center
        sc.number_format = "0.0000"

    # ===== 区域3：汇总 =====
    row += 2
    ws.cell(row=row, column=1, value="③ 单个成本汇总").font = Font(bold=True, size=12)
    row += 1
    for i, spec in enumerate(specs):
        c = ws.cell(row=row, column=1 + i, value=spec)
        c.font = header_font; c.fill = header_fill; c.border = border; c.alignment = center
    row += 1
    for i, spec in enumerate(specs):
        cost_col = 2 + i + len(specs)
        c = ws.cell(row=row, column=1 + i, value=f'={col_letter(cost_col)}{total_row}')
        c.fill = calc_fill; c.border = border; c.alignment = center
        c.number_format = "0.0000"

    # 列宽
    ws.column_dimensions["A"].width = 16
    for i in range(len(specs) * 2):
        ws.column_dimensions[col_letter(2 + i)].width = 13
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def num(v):
    """字符串转数字，空/非法返回 None"""
    if v is None:
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return None


def col_letter(n):
    """1->A, 27->AA"""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def get_col_letter(n):
    return col_letter(n)


# ---------- HTTP 服务 ----------

class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=BASE_DIR, **kwargs)

    def _send_json(self, obj, code=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, data, filename, ctype):
        from urllib.parse import quote
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        # 中文文件名用 RFC 5987 编码，避免 latin-1 报错
        encoded = quote(filename)
        self.send_header("Content-Disposition",
                         f"attachment; filename=\"{encoded}\"; filename*=UTF-8''{encoded}")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        path = self.path.split("?")[0]
        if path == "/api/materials":
            self._send_json(load_materials())
            return
        if path == "/api/products":
            self._send_json(load_products())
            return
        if path == "/api/export":
            # /api/export?product=<key>
            qs = self.path.split("?", 1)[1] if "?" in self.path else ""
            pk = ""
            for part in qs.split("&"):
                if part.startswith("product="):
                    pk = part[8:]
            products = load_products()
            product = products.get(pk)
            if not product:
                self._send_json({"error": "product not found"}, 404)
                return
            data = build_excel(load_materials(), product)
            safe_name = re.sub(r'[^\w\u4e00-\u9fff-]', '_', product.get("name", "product"))
            self._send_file(data, f"{safe_name}成本表.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            return
        super().do_GET()

    def do_POST(self):
        path = self.path.split("?")[0]
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length)
        try:
            data = json.loads(raw.decode("utf-8"))
        except json.JSONDecodeError:
            self._send_json({"error": "invalid json"}, 400)
            return

        if path == "/api/materials":
            if not isinstance(data, dict):
                self._send_json({"error": "must be object"}, 400)
                return
            save_json(MATERIALS_FILE, data)
            self._send_json({"ok": True})
            return
        if path == "/api/products":
            if not isinstance(data, dict):
                self._send_json({"error": "must be object"}, 400)
                return
            save_json(PRODUCTS_FILE, data)
            self._send_json({"ok": True})
            return
        self._send_json({"error": "not found"}, 404)

    def log_message(self, fmt, *args):
        pass  # 静默


def main():
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("0.0.0.0", PORT), Handler) as httpd:
        print(f"food-cost server on port {PORT}")
        httpd.serve_forever()


if __name__ == "__main__":
    main()
